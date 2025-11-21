# app.py — Part 1 of 4
import streamlit as st
import pandas as pd
import re
import io
from openpyxl import load_workbook

st.set_page_config(
    page_title="KLD Excel → SVG Generator (Grey-detect, Seals)",
    layout="wide"
)

st.title("📏 KLD Excel → SVG Generator (Grey-detect + Seals + Strict Validation)")
st.caption("Detects grey header region, extracts header until numeric table, applies strict dimension validation, and generates SVG dieline.")

# ===========================================
# Helpers (Clean, Correct)
# ===========================================

def clean_numeric_list(seq):
    """
    Converts a list of Excel cell values into a list of floats.
    Removes blanks, 'nan', None, etc.
    Extracts embedded numbers inside text if needed.
    """
    out = []
    for v in seq:
        s = str(v).strip().replace(",", "")
        if not s or s.lower() in ("nan", "none"):
            continue
        try:
            out.append(float(s))
        except:
            m = re.search(r"(-?\d+(?:\.\d+)?)", s)
            if m:
                out.append(float(m.group(1)))
    return out


def first_pair_from_text(text):
    """
    Extracts dimension pairs like '416 * 386' or '416x386'.
    Returns (width_mm, cut_length_mm)
    """
    text = str(text)
    m = re.search(r"(\d+(?:\.\d+)?)\s*[\*xX]\s*(\d+(?:\.\d+)?)", text)
    if m:
        return float(m.group(1)), float(m.group(2))
    return 0, 0


# ===========================================
# Grey detection (kept as before)
# ===========================================
def cell_is_filled(cell):
    """
    Returns True if the cell has a visible non-white fill.
    """
    try:
        fl = getattr(cell, "fill", None)
        if not fl:
            return False

        patt = getattr(fl, "patternType", None)
        if patt is None or str(patt).lower() == "none":
            return False

        fg = getattr(fl, "fgColor", None)
        if not fg:
            return False

        rgb = getattr(fg, "rgb", None)
        if rgb:
            rgb = str(rgb).upper()
            if rgb in ("FFFFFFFF", "FFFFFF", "00FFFFFF"):
                return False
            return True

        if getattr(fg, "indexed", None) is not None or getattr(fg, "theme", None) is not None:
            return True

        return False

    except Exception:
        return False


# ===========================================
# Extraction (beginning)
# Part 2 will include numeric-table scanning & sequences
# ===========================================
def extract_kld_data_from_bytes(xl_bytes):
    bytes_io = io.BytesIO(xl_bytes)

    # Load workbook and worksheet
    wb = load_workbook(bytes_io, data_only=True, read_only=False)
    ws = wb.active

    # Detect filled positions (grey header)
    filled_positions = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            try:
                if cell_is_filled(ws.cell(r, c)):
                    filled_positions.append((r, c))
            except:
                continue

    # Determine grey region bounds (if any)
    if filled_positions:
        rows = [r for r, _ in filled_positions]
        cols = [c for _, c in filled_positions]
        r_min, r_max = min(rows), max(rows)
        c_min, c_max = min(cols), max(cols)
    else:
        # fallback defaults
        r_min, r_max, c_min, c_max = 2, 68, 2, 28
        r_max = min(r_max, ws.max_row)
        c_max = min(c_max, ws.max_column)

    # Build grey cols by row map
    grey_cols_by_row = {}
    for r, c in filled_positions:
        grey_cols_by_row.setdefault(r, set()).add(c)

    # Header extraction using grey columns
    header_lines = []
    detected_dim_line = ""

    if grey_cols_by_row:
        rows_to_scan = sorted(r for r in grey_cols_by_row.keys() if r_min <= r <= r_max)
    else:
        rows_to_scan = list(range(r_min, r_max + 1))

    for r in rows_to_scan:
        row_vals = []
        numeric_count = 0

        # Use only grey columns for header extraction
        if r in grey_cols_by_row:
            cols = sorted(grey_cols_by_row[r])
        else:
            cols = list(range(c_min, c_max + 1))

        for c in cols:
            val = ws.cell(r, c).value
            if val is None:
                continue
            sval = str(val).strip()
            if not sval:
                continue

            if re.match(r"^-?\d+(\.\d+)?$", sval):
                numeric_count += 1

            row_vals.append(sval)

        # Stop header collection when numeric table begins
        if numeric_count >= 3:
            break

        line_text = " ".join(row_vals).strip()
        if line_text:
            header_lines.append(line_text)
            if not detected_dim_line and re.search(r"dimension|width|cut", line_text, re.IGNORECASE):
                detected_dim_line = line_text

    # Dimension extraction
    if detected_dim_line:
        w_val, c_val = first_pair_from_text(detected_dim_line)
    else:
        w_val, c_val = 0, 0
        for ln in header_lines:
            t1, t2 = first_pair_from_text(ln)
            if t1 and t2:
                w_val, c_val = t1, t2
                break

    width_mm = int(w_val) if w_val else 0
    cut_length_mm = int(c_val) if c_val else 0

    dimension_text = f"Dimension ( Width * Cut-off length ) : {width_mm} * {cut_length_mm} ( in mm )"

    job_name = next((ln for ln in header_lines if re.search(r"[A-Za-z]", ln)), "KLD Layout")

    # --- Prepare for numeric-table scanning (Parts 2..4 will follow) ---
    # Keep workbook (ws), df_orig (pandas with original indices), start_row_index, etc.
    bytes_io.seek(0)
    df_orig = pd.read_excel(bytes_io, header=None, engine="openpyxl")
    df_orig = df_orig.fillna("").astype(str)
    mask = df_orig.apply(lambda r: any(str(x).strip() for x in r), axis=1)
    df = df_orig[mask]

    # find numeric-table start row (index in original df)
    start_row_index = None
    for idx in df.index[:min(120, len(df))]:
        numeric_count = sum(1 for c in df.loc[idx].tolist() if re.match(r"^\d+(\.\d+)?$", str(c).strip()))
        if numeric_count >= 3:
            start_row_index = idx
            break
    if start_row_index is None:
        start_row_index = df.index[0] if len(df.index) else 0

    # df_num preserves original Excel indices
    df_num = df.loc[start_row_index:].copy()

    # Now we hand off to Part 2 which implements top_seq & side_seq scanning
    # (I'll send Part 2 next if you want)
    return {
        "job_name": job_name,
        "header_lines": header_lines,
        "dimension_text": dimension_text,
        "width_mm": width_mm,
        "cut_length_mm": cut_length_mm,
        # placeholders until Part 2 fills actual sequences
        "top_seq": "",
        "side_seq": "",
    }
# ============================================================
# PART 2 — NUMERIC TABLE SCANNING (side_seq + top_seq)
# ============================================================

def extract_sequences(ws, df_num, start_row_index, width_mm, cut_length_mm):
    """
    Returns: (top_seq_list, side_seq_list)
    Using merged/unmerged logic, row/col gap logic, scanning entire sheet (Option D).
    """

    # --------------------------------------------------------
    # 1. Detect merged ranges in the worksheet
    # --------------------------------------------------------
    merged_ranges = []
    for rng in ws.merged_cells.ranges:
        merged_ranges.append((rng.min_row, rng.max_row, rng.min_col, rng.max_col))

    def find_merged_block(r, c):
        """Return (r1,r2,c1,c2) if cell is inside a merged block."""
        for (r1, r2, c1, c2) in merged_ranges:
            if r1 <= r <= r2 and c1 <= c <= c2:
                return (r1, r2, c1, c2)
        return None

    def extract_block_value(r1, r2, c1, c2):
        """Return single numeric value from merged block (must be consistent)."""
        vals = []
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                v = ws.cell(rr, cc).value
                if v is None:
                    continue
                sval = str(v).strip()
                if re.match(r"^-?\d+(?:\.\d+)?$", sval):
                    vals.append(float(sval))

        if not vals:
            return None
        if len(set(vals)) > 1:
            # conflicting values inside merged block → not allowed
            raise ValueError(
                f"Merged cell block {r1}-{r2} {c1}-{c2} contains multiple numeric values: {vals}"
            )
        return vals[0]

    # --------------------------------------------------------
    # 2. Identify qualifying vertical column for side_seq
    # --------------------------------------------------------
    col_numeric_counts = {}
    for c in df_num.columns:
        count = 0
        for v in df_num[c].tolist():
            sval = str(v).strip()
            if re.match(r"^-?\d+(?:\.\d+)?$", sval):
                count += 1
        col_numeric_counts[c] = count

    # Select first column (left-most) with >= 2 numeric values
    side_col = None
    for c in df_num.columns:
        if col_numeric_counts[c] >= 2:
            side_col = c
            break

    if side_col is None:
        return [], []  # No valid numeric columns found

    excel_col = side_col + 1  # convert df column to Excel column index

    # --------------------------------------------------------
    # 3. Collect numeric blocks vertically (merged + unmerged)
    # --------------------------------------------------------
    collected = []  # list of (excel_row, numeric_value)
    seen_blocks = set()

    for df_idx in df_num.index:
        excel_row = df_idx + 1  # df index → Excel row number
        val = df_num.loc[df_idx, side_col]
        sval = str(val).strip()

        blk = find_merged_block(excel_row, excel_col)

        if blk:
            (r1, r2, c1, c2) = blk

            # Avoid processing merged block twice
            if (r1, r2, c1, c2) in seen_blocks:
                continue
            seen_blocks.add((r1, r2, c1, c2))

            block_val = extract_block_value(r1, r2, c1, c2)
            if block_val is not None:
                collected.append((r1, block_val))

        else:
            # unmerged numeric cell
            if re.match(r"^-?\d+(?:\.\d+)?$", sval):
                collected.append((excel_row, float(sval)))

    # --------------------------------------------------------
    # 4. Sort numeric blocks by row order
    # --------------------------------------------------------
    collected.sort(key=lambda x: x[0])

    # --------------------------------------------------------
    # 5. Apply merged/unmerged gap-logic (YOUR FINAL RULE)
    # --------------------------------------------------------
    side_seq = []
    last_unmerged_row = None

    for (row, value) in collected:
        blk = find_merged_block(row, excel_col)

        if blk:
            # merged → always include
            side_seq.append(value)
            continue

        # unmerged → apply gap logic
        if last_unmerged_row is None:
            side_seq.append(value)
            last_unmerged_row = row
            continue

        gap = row - last_unmerged_row

        if gap <= 2:  # ≤ 1 blank row allowed → gap 2 (1 blank) is OK
            side_seq.append(value)
            last_unmerged_row = row
        else:
            break  # STOP: more than 1 blank between unmerged values

    # --------------------------------------------------------
    # 6. TOP SEQUENCE (horizontal scanning mirror logic)
    # --------------------------------------------------------
    # find first row with >=2 numerics
    row_numeric_counts = {}
    for df_row in df_num.index:
        row_list = df_num.loc[df_row].tolist()
        cnt = sum(1 for v in row_list if re.match(r"^-?\d+(?:\.\d+)?$", str(v).strip()))
        row_numeric_counts[df_row] = cnt

    top_row = None
    for r in df_num.index:
        if row_numeric_counts[r] >= 2:
            top_row = r
            break

    if top_row is None:
        return [], side_seq  # no usable top row

    # extract horizontally
    collected_h = []
    seen_blocks_h = set()

    for c in df_num.columns:
        excel_col = c + 1
        excel_row = top_row + 1

        val = df_num.loc[top_row, c]
        sval = str(val).strip()

        blk = find_merged_block(excel_row, excel_col)

        if blk:
            (r1, r2, c1, c2) = blk
            if (r1, r2, c1, c2) in seen_blocks_h:
                continue
            seen_blocks_h.add((r1, r2, c1, c2))

            block_val = extract_block_value(r1, r2, c1, c2)
            if block_val is not None:
                collected_h.append((c1, block_val))  # use leftmost column as position
        else:
            if re.match(r"^-?\d+(?:\.\d+)?$", sval):
                collected_h.append((excel_col, float(sval)))

    # sort horizontally by column number
    collected_h.sort(key=lambda x: x[0])

    # apply horizontal gap logic
    top_seq = []
    last_unmerged_col = None

    for (col, value) in collected_h:
        blk = find_merged_block(top_row + 1, col)

        if blk:
            top_seq.append(value)
            continue

        if last_unmerged_col is None:
            top_seq.append(value)
            last_unmerged_col = col
            continue

        gap = col - last_unmerged_col

        if gap <= 2:  # at most one blank column
            top_seq.append(value)
            last_unmerged_col = col
        else:
            break  # STOP

    return top_seq, side_seq
# ============================================================
# PART 3 — Integrate sequence extraction into Part 1 function
# ============================================================

def extract_kld_data_from_bytes(xl_bytes):
    bytes_io = io.BytesIO(xl_bytes)

    # Load workbook and worksheet
    wb = load_workbook(bytes_io, data_only=True, read_only=False)
    ws = wb.active

    # -----------------------------------------------
    # GREY HEADER EXTRACTION (same as Part 1)
    # -----------------------------------------------
    filled_positions = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            try:
                if cell_is_filled(ws.cell(r, c)):
                    filled_positions.append((r, c))
            except:
                continue

    if filled_positions:
        rows = [r for r, _ in filled_positions]
        cols = [c for _, c in filled_positions]
        r_min, r_max = min(rows), max(rows)
        c_min, c_max = min(cols), max(cols)
    else:
        r_min, r_max, c_min, c_max = 2, 68, 2, 28
        r_max = min(r_max, ws.max_row)
        c_max = min(c_max, ws.max_column)

    grey_cols_by_row = {}
    for r, c in filled_positions:
        grey_cols_by_row.setdefault(r, set()).add(c)

    header_lines = []
    detected_dim_line = ""

    if grey_cols_by_row:
        rows_to_scan = sorted(r for r in grey_cols_by_row.keys() if r_min <= r <= r_max)
    else:
        rows_to_scan = list(range(r_min, r_max + 1))

    for r in rows_to_scan:
        row_vals = []
        numeric_count = 0

        cols = sorted(grey_cols_by_row[r]) if r in grey_cols_by_row else list(range(c_min, c_max + 1))

        for c in cols:
            val = ws.cell(r, c).value
            if val is None:
                continue
            sval = str(val).strip()
            if not sval:
                continue

            if re.match(r"^-?\d+(?:\.\d+)?$", sval):
                numeric_count += 1

            row_vals.append(sval)

        if numeric_count >= 3:
            break

        line_text = " ".join(row_vals).strip()
        if line_text:
            header_lines.append(line_text)
            if not detected_dim_line and re.search(r"dimension|width|cut", line_text, re.IGNORECASE):
                detected_dim_line = line_text

    if detected_dim_line:
        w_val, c_val = first_pair_from_text(detected_dim_line)
    else:
        w_val, c_val = 0, 0
        for ln in header_lines:
            t1, t2 = first_pair_from_text(ln)
            if t1 and t2:
                w_val, c_val = t1, t2
                break

    width_mm = int(w_val) if w_val else 0
    cut_length_mm = int(c_val) if c_val else 0

    dimension_text = f"Dimension ( Width * Cut-off length ) : {width_mm} * {cut_length_mm} ( in mm )"
    job_name = next((ln for ln in header_lines if re.search(r"[A-Za-z]", ln)), "KLD Layout")

    # -----------------------------------------------
    # LOAD WHOLE SHEET FOR NUMERIC TABLE (OPTION D)
    # -----------------------------------------------
    bytes_io.seek(0)
    df_orig = pd.read_excel(bytes_io, header=None, engine="openpyxl")
    df_orig = df_orig.fillna("").astype(str)

    mask = df_orig.apply(lambda r: any(str(x).strip() for x in r), axis=1)
    df = df_orig[mask]

    start_row_index = None
    for idx in df.index[: min(120, len(df))]:
        numeric_count = sum(1 for c in df.loc[idx].tolist() if re.match(r"^\d+(\.\d+)?$", str(c).strip()))
        if numeric_count >= 3:
            start_row_index = idx
            break

    if start_row_index is None:
        start_row_index = df.index[0]

    df_num = df.loc[start_row_index:].copy()

    # -----------------------------------------------
    # PART 2: Extract sequences
    # -----------------------------------------------
    top_seq_list, side_seq_list = extract_sequences(
        ws,           # worksheet
        df_num,       # numeric-table dataframe
        start_row_index,
        width_mm,
        cut_length_mm
    )

    # -----------------------------------------------
    # FORMAT LISTS AS CSV STRINGS
    # -----------------------------------------------
    def fmt_list(lst):
        out = []
        for v in lst:
            s = str(v)
            if "." in s:
                s = s.rstrip("0").rstrip(".")
            out.append(s)
        return ",".join(out)

    top_seq = fmt_list(top_seq_list)
    side_seq = fmt_list(side_seq_list)

    # -----------------------------------------------
    # FINAL RETURN STRUCTURE
    # -----------------------------------------------
    return {
        "job_name": job_name,
        "header_lines": header_lines,
        "dimension_text": dimension_text,
        "width_mm": width_mm,
        "cut_length_mm": cut_length_mm,
        "top_seq": top_seq,
        "side_seq": side_seq,
    }
# ============================================================
# PART 4 — SVG generator + Streamlit UI
# ============================================================

def make_svg(data, line_spacing_mm=5.0):

    import re

    def parse_seq(src):
        if not src:
            return []
        parts = re.split(r"[,;|]", str(src))
        return [
            float(p.strip())
            for p in parts
            if re.match(r"^-?\d+(?:\.\d+)?$", p.strip())
        ]

    W = float(data["cut_length_mm"]) if data.get("cut_length_mm") else 0.0
    H = float(data["width_mm"]) if data.get("width_mm") else 0.0
    top_seq = parse_seq(data["top_seq"])
    side_seq = parse_seq(data["side_seq"])

    header_lines = data.get("header_lines") or []
    if not header_lines:
        header_lines = [
            data.get("job_name", "KLD Layout"),
            data.get("dimension_text", "")
        ]

    # Artboard expansion
    extra = 60
    margin = extra / 2
    canvas_W = W + extra
    canvas_H = H + extra

    dieline = "#92278f"
    stroke_pt = 0.356
    font_mm = 0.8
    tick_short = 5
    top_shift_up = 5
    left_shift_left = 5
    left_text_shift_right = 6
    top_text_shift_down = 4
    crop_off = 5
    crop_len = 5

    out = []
    out.append(
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{canvas_W}mm" height="{canvas_H}mm" '
        f'viewBox="0 0 {canvas_W} {canvas_H}">'
    )
    out.append('<defs><style><![CDATA[')
    out.append(f'.dieline{{stroke:{dieline};stroke-width:{stroke_pt}pt;fill:none;}}')
    out.append(f'.text{{font-family:Arial; font-size:{font_mm}mm; fill:{dieline};}}')
    out.append(']]></style></defs>')

    # Header block
    out.append('<g id="Header">')
    cx = canvas_W / 2
    for i, line in enumerate(header_lines):
        y = (i + 1) * line_spacing_mm
        out.append(
            f'<text x="{cx}" y="{y}" text-anchor="middle" class="text">{line}</text>'
        )
    out.append('</g>')

    # Dieline rectangle
    out.append(
        f'<rect x="{margin}" y="{margin}" width="{W}" height="{H}" class="dieline"/>'
    )

    # Measurements
    out.append('<g id="Measurements">')

    # Top ticks
    x = 0
    out.append(
        f'<line x1="{margin}" y1="{margin-top_shift_up}" '
        f'x2="{margin}" y2="{margin-top_shift_up-tick_short}" class="dieline"/>'
    )

    for v in top_seq:
        x += v
        out.append(
            f'<line x1="{margin+x}" y1="{margin-top_shift_up}" '
            f'x2="{margin+x}" y2="{margin-top_shift_up-tick_short}" class="dieline"/>'
        )
        mid = x - v / 2
        out.append(
            f'<text x="{margin+mid}" y="{margin-top_shift_up-tick_short-1+top_text_shift_down}" '
            f'text-anchor="middle" class="text">{round(v, 2)}</text>'
        )

    # Side ticks
    y = 0
    out.append(
        f'<line x1="{margin-left_shift_left}" y1="{margin}" '
        f'x2="{margin-left_shift_left-tick_short}" y2="{margin}" class="dieline"/>'
    )

    for v in side_seq:
        y += v
        out.append(
            f'<line x1="{margin-left_shift_left}" y1="{margin+y}" '
            f'x2="{margin-left_shift_left-tick_short}" y2="{margin+y}" class="dieline"/>'
        )
        midY = y - v / 2
        lx = margin - left_shift_left - tick_short - 2 + left_text_shift_right
        out.append(
            f'<text x="{lx}" y="{margin+midY}" '
            f'transform="rotate(-90 {lx} {margin+midY})" '
            f'text-anchor="middle" class="text">{round(v, 2)}</text>'
        )

    out.append('</g>')

    # Crop marks
    out.append('<g id="CropMarks">')
    out.append(
        f'<line x1="{margin+W+crop_off}" y1="{margin}" '
        f'x2="{margin+W+crop_off+crop_len}" y2="{margin}" class="dieline"/>'
    )
    out.append(
        f'<line x1="{margin+W+crop_off}" y1="{margin+H}" '
        f'x2="{margin+W+crop_off+crop_len}" y2="{margin+H}" class="dieline"/>'
    )
    out.append('</g>')

    # Photocell mark
    photocell_w, photocell_h = 6, 12
    pc_x = margin + W - photocell_w
    pc_y = margin

    out.append('<g id="Photocell">')
    out.append(
        f'<rect x="{pc_x}" y="{pc_y}" width="{photocell_w}" '
        f'height="{photocell_h}" class="dieline"/>'
    )
    out.append(
        f'<line x1="{pc_x+photocell_w}" y1="{pc_y}" '
        f'x2="{pc_x+photocell_w+3}" y2="{pc_y-3}" class="dieline"/>'
    )
    out.append(
        f'<text x="{pc_x+photocell_w+2}" y="{pc_y-4}" class="text">'
        f'Photocell Mark {photocell_w}×{photocell_h} mm</text>'
    )
    out.append('</g>')

    # Width marker
    total_side = sum(side_seq) if side_seq else 0
    wx = pc_x + photocell_w + 4
    midY = margin + total_side / 2

    out.append('<g id="WidthMarker">')
    out.append(
        f'<line x1="{wx}" y1="{margin}" x2="{wx}" y2="{margin+total_side}" '
        f'class="dieline"/>'
    )
    out.append(
        f'<text x="{wx+6}" y="{midY}" transform="rotate(-90 {wx+6} {midY})" '
        f'class="text" text-anchor="middle">width = {round(total_side,2)} mm</text>'
    )
    out.append('</g>')

    # Height marker
    total_top = sum(top_seq) if top_seq else 0
    hy = margin + total_side + 5

    out.append('<g id="HeightMarker">')
    out.append(
        f'<line x1="{margin}" y1="{hy}" x2="{margin+total_top}" y2="{hy}" '
        f'class="dieline"/>'
    )
    out.append(
        f'<text x="{margin+total_top/2}" y="{hy+6}" text-anchor="middle" '
        f'class="text">height = {round(total_top,2)} mm</text>'
    )
    out.append('</g>')

    # Dynamic boxes
    out.append('<g id="DynamicBoxes">')
    if top_seq and side_seq:
        max_top = max(top_seq)
        max_idx = top_seq.index(max_top)
        left_x = margin + sum(top_seq[:max_idx])
        skip = 2
        while skip < len(side_seq):
            top_y = margin + sum(side_seq[:skip])
            height = side_seq[skip]
            out.append(
                f'<rect x="{left_x}" y="{top_y}" width="{max_top}" '
                f'height="{height}" class="dieline"/>'
            )
            skip += 3
    out.append('</g>')

    # Seals
    out.append('<g id="Seals">')

    first_top = top_seq[0] if top_seq else 0
    last_top = top_seq[-1] if top_seq else 0

    left_end_x = margin + first_top / 2
    out.append(
        f'<text x="{left_end_x}" y="{midY}" transform="rotate(-90 {left_end_x} {midY})" '
        f'text-anchor="middle" class="text">END SEAL</text>'
    )

    right_end_x = margin + W - last_top/2
    out.append(
        f'<text x="{right_end_x}" y="{midY}" transform="rotate(90 {right_end_x} {midY})" '
        f'text-anchor="middle" class="text">END SEAL</text>'
    )

    mid_top_x = margin + total_top / 2
    first_side = side_seq[0] if side_seq else 0
    last_side = side_seq[-1] if side_seq else 0

    top_center_y = margin + first_side / 2
    bottom_center_y = margin + total_side - last_side / 2

    out.append(
        f'<text x="{mid_top_x}" y="{top_center_y}" text-anchor="middle" '
        f'class="text">CENTER SEAL</text>'
    )
    out.append(
        f'<text x="{mid_top_x}" y="{bottom_center_y}" text-anchor="middle" '
        f'class="text">CENTER SEAL</text>'
    )

    out.append('</g>')
    out.append('</svg>')

    return "\n".join(out)


# ============================================================
# STREAMLIT APP
# ============================================================

uploaded_file = st.file_uploader("Upload KLD Excel file", type=["xlsx", "xls"])

if uploaded_file:
    try:
        raw = uploaded_file.read()

        data = extract_kld_data_from_bytes(raw)

        # Parse sequences for validation
        def parse_seq_list(src):
            import re
            parts = re.split(r"[,;|]", str(src))
            return [
                float(p.strip())
                for p in parts
                if re.match(r"^-?\d+(?:\.\d+)?$", p.strip())
            ]

        top_seq = parse_seq_list(data["top_seq"])
        side_seq = parse_seq_list(data["side_seq"])

        sum_top = sum(top_seq)
        sum_side = sum(side_seq)

        cut_length_mm = data["cut_length_mm"]
        width_mm = data["width_mm"]

        # STRICT VALIDATION
        errors = []

        if cut_length_mm and sum_top and abs(sum_top - float(cut_length_mm)) > 0.001:
            errors.append(
                f"Sum of top_seq = {sum_top} mm does NOT match cut_length_mm = {cut_length_mm} mm."
            )

        if width_mm and sum_side and abs(sum_side - float(width_mm)) > 0.001:
            errors.append(
                f"Sum of side_seq = {sum_side} mm does NOT match width_mm = {width_mm} mm."
            )

        if errors:
            st.error("❌ SVG generation blocked due to mismatched dimensions:")
            for e in errors:
                st.write(f"- {e}")
            st.stop()

        svg = make_svg(data)
        st.success("✅ Dimensions validated. No mismatches detected.")

        st.download_button(
            "⬇️ Download SVG File",
            svg,
            f"{uploaded_file.name}_layout.svg",
            "image/svg+xml"
        )

        st.code(f"side_seq: {data['side_seq']}\ntop_seq: {data['top_seq']}", language="text")

    except Exception as e:
        st.error(f"❌ Conversion failed: {e}")

else:
    st.info("Please upload the Excel (.xlsx) file to begin.")
